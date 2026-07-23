import os
import csv
import openpyxl
from flask import Flask, request, send_file

app = Flask(__name__)

@app.route('/upload-and-process', methods=['POST'])
def upload_and_process():
    if 'file' not in request.files:
        return {'error': 'Không tìm thấy file gửi lên'}, 400
    
    file = request.files['file']
    if file.filename == '':
        return {'error': 'Chưa chọn file'}, 400

    filename = file.filename
    name, ext = os.path.splitext(filename)
    input_path = os.path.join('/tmp', filename)
    
    # Tên file Excel kết quả trả về tương ứng với tên file dữ liệu gốc
    output_filename = f"Xu_ly_ket_qua_{name}.xlsx"
    output_path = os.path.join('/tmp', output_filename)

    file.save(input_path)

    try:
        with open(input_path, 'r', encoding='utf-8-sig', errors='ignore') as f:
            sample_text = f.read(2048)
        
        delimiter = ','
        if '\t' in sample_text:
            delimiter = '\t'
        elif ';' in sample_text:
            delimiter = ';'

        rows = []
        encodings = ['utf-8-sig', 'utf-16le', 'utf-16', 'latin-1']
        success = False
        for enc in encodings:
            try:
                with open(input_path, 'r', encoding=enc) as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    rows = list(reader)
                if rows:
                    success = True
                    break
            except Exception:
                continue

        if not success or len(rows) < 3:
            return {'error': 'File rỗng hoặc ít hơn 3 dòng.'}, 400

        # --- BƯỚC 1: XÂY DỰNG CẤU TRÚC BẢNG (THÊM HÀNG VÀ CỘT) ---
        processed_rows = []

        # 1. Xử lý 3 dòng tiêu đề đầu tiên (Chèn 2 cột rỗng vào index 1 và index 3)
        for i in range(min(3, len(rows))):
            r = list(rows[i])
            while len(r) <= 1:
                r.append("")
            r.insert(1, "")
            r.insert(3, "")
            processed_rows.append(r)

        # 2. Tạo dòng thứ 4: Đếm từ 1 -> 81 bắt đầu từ ô C4
        sample_r = list(rows[2]) if len(rows) > 2 else []
        total_cols = len(sample_r) + 2
        row4 = [""] * total_cols
        count = 1
        for col_idx in range(2, len(row4)):
            if count <= 81:
                row4[col_idx] = str(count)
                count += 1
        processed_rows.append(row4)

        # 3. Xử lý các dòng dữ liệu từ dòng 5 trở đi (thêm STT)
        stt = 1
        for i in range(3, len(rows)):
            if not rows[i] or not any(rows[i]):
                continue
            r = list(rows[i])
            r.insert(1, str(stt))
            r.insert(3, str(stt))
            processed_rows.append(r)
            stt += 1

        # --- BƯỚC 2: ĐIỀN CÔNG THỨC VÀO ĐÚNG CỘT UN ---
        header = processed_rows[2] if len(processed_rows) > 2 else []
        un_idx = -1
        for idx, val in enumerate(header):
            if val.strip().upper() == 'UN':
                un_idx = idx
                break

        if un_idx != -1:
            for i in range(4, len(processed_rows)):
                r = processed_rows[i]
                if not r or not any(r):
                    continue
                if len(r) > un_idx:
                    row_num = i + 1
                    formula = f"=MAX(ABS(E{row_num}-H{row_num}), ABS(F{row_num}-H{row_num}), ABS(G{row_num}-H{row_num}))/H{row_num}*100"
                    r[un_idx] = formula

        # --- BƯỚC 3: TỰ ĐỘNG TÌM FILE TEMPLATE EXCEL TƯƠNG ỨNG ---
        # Danh sách các kiểu tên file template có thể khớp với tên file dữ liệu
        possible_templates = [
            f"Xu ly du lieu {name}.xlsx",
            f"Xu ly du lieu {name.replace('_', ' ')}.xlsx",
            f"{name}.xlsx",
            "Xu ly du lieu file do.xlsx"  # File mặc định fallback cuối cùng
        ]
        
        template_excel = "Xu ly du lieu file do.xlsx"
        for t in possible_templates:
            if os.path.exists(t):
                template_excel = t
                break

        wb = openpyxl.load_workbook(template_excel)

        sheet_name = 'Ket qua do'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # Xóa dữ liệu cũ trên sheet này trước khi ghi mới
        ws.delete_rows(1, ws.max_row + 1)

        # Ghi dữ liệu dòng/cột vào worksheet
        for r_idx, row in enumerate(processed_rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                if isinstance(val, str) and val.startswith('='):
                    ws.cell(row=r_idx, column=c_idx, value=val)
                else:
                    try:
                        if val == '':
                            cell_val = None
                        elif '.' in val or 'e' in val.lower():
                            cell_val = float(val)
                        else:
                            cell_val = int(val)
                    except ValueError:
                        cell_val = val
                    ws.cell(row=r_idx, column=c_idx, value=cell_val)

        wb.save(output_path)

        if os.path.exists(input_path):
            os.remove(input_path)

        return send_file(
            output_path, 
            as_attachment=True, 
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return {'error': f"Lỗi xử lý file: {str(e)}"}, 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
