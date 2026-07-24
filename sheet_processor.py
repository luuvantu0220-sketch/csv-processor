import os
import pandas as pd
import openpyxl

def process_with_embedded_template(datasource_path, output_path):
    # Đọc file nguồn với dtype=str để bảo toàn dữ liệu và tiết kiệm RAM
    if datasource_path.endswith('.csv'):
        try:
            df_source = pd.read_csv(datasource_path, encoding='utf-8-sig', dtype=str, low_memory=False)
        except Exception:
            df_source = pd.read_csv(datasource_path, encoding='latin-1', errors='ignore', dtype=str, low_memory=False)
    else:
        df_source = pd.read_excel(datasource_path, sheet_name=0, dtype=str)
    
    # Xác định đường dẫn tuyệt đối của file mẫu dựa vào vị trí file script hiện tại
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, 'Xu ly du lieu file do.xlsx')
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Không tìm thấy file mẫu tại đường dẫn: {template_path}")
        
    # Load workbook
    wb_template = openpyxl.load_workbook(template_path)
    
    if 'Ket qua do' in wb_template.sheetnames:
        ws_result = wb_template['Ket qua do']
    else:
        ws_result = wb_template.create_sheet('Ket qua do')
        
    # Xóa dữ liệu cũ trong sheet kết quả (nếu có) để tránh bị chồng chéo
    if ws_result.max_row > 0:
        ws_result.delete_rows(1, ws_result.max_row + 1)
        
    # Ghi Header trước
    headers = list(df_source.columns)
    ws_result.append(headers)

    # Ghi toàn bộ dữ liệu hàng loạt xuống sheet 'Ket qua do'
    rows_data = df_source.values.tolist()
    for row in rows_data:
        ws_result.append(row)
            
    wb_template.save(output_path)
    return output_path
